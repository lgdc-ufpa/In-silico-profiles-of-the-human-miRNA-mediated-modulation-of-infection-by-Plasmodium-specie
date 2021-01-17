import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from openpyxl import load_workbook, Workbook
import os

#TODO: get miRNAs
list_mirnas = []
list_species = []
with open('output/df_mirnas.csv', 'r') as f:
    for elem in f:
        mirna = elem.replace("\n", "")
        list_mirnas.append(mirna)

print(list_mirnas)

#TODO: get species
book = load_workbook('output/df_google_sheets_sankey_formated.xlsx')
sheet = book.active
index = 2
while not sheet['B'+str(index)].value == None:
    specie = sheet['B' + str(index)].value
    value = sheet['C' +  str(index)].value
    print(specie)
    if specie not in list_species:
        list_species.append(specie)
    index += 1
book.close()

print("Species\n")
for specie in list_species:
    print(specie)

# #TODO: create matrix
# book = Workbook()
# cwd = os.getcwd()
# book.save(os.path.join(cwd, 'output', 'df_heatmap.xlsx'))
# book.close()

#TODO: fill matrix
# book_heatmap = load_workbook(os.path.join('output', 'df_heatmap.xlsx'))
# sheet_heatmap = book.active

book = load_workbook(os.path.join('output', 'df_google_sheets_sankey_formated.xlsx'))
sheet = book.active

index = 2
while not sheet['A'+str(index)].value == None:
    mirna = sheet['A'+str(index)].value
    specie = sheet['B'+str(index)].value
    value = sheet['C'+str(index)].value

    print(mirna, specie, value)
    index += 1
